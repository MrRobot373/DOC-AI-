"""
Kimi-style engineering document analyzer.

This is intentionally separate from the existing Flask app. It is a CLI-first
experiment for deeper document review and stricter finding validation.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from doc_parser import get_document_summary, get_section_chunks, parse_document  # noqa: E402
from ollama import Client  # noqa: E402


CATEGORIES = {
    "Grammar & Spelling": "📝",
    "Terminology Consistency": "📖",
    "Flowchart ↔ Description": "🔄",
    "Units & Calculations": "🔬",
    "Formatting & Alignment": "📐",
    "Signal & Variable Naming": "🏷️",
    "Test Result Completeness": "✅",
    "Waveform Documentation": "📈",
    "Cross-Reference Accuracy": "🔗",
    "Logical Consistency": "🧠",
    "Connector & Pin Mapping": "🔌",
    "Measurement Resolution": "🔍",
    "Decimal Digit Consistency": "🔢",
    "Table Quality": "📊",
    "Subscript/Superscript": "⬇️",
    "Datasheet Copy Error": "📋",
    "TOC & Heading Structure": "📚",
}

SEVERITIES = {"High", "Medium", "Low"}
SEVERITY_ORDER = {"High": 0, "Medium": 1, "Low": 2}

NON_ERROR_PHRASES = [
    "actually correct",
    "is correct",
    "spelled correctly",
    "not found",
    "acceptable",
    "adequate",
    "consistent between",
    "this is actually correct",
    "no issue",
    "not an issue",
]


@dataclass
class Finding:
    category: str
    description: str
    location: str
    severity: str
    details: str
    evidence: str = ""
    source: str = "local"

    @property
    def icon(self) -> str:
        return CATEGORIES.get(self.category, "")


def main() -> int:
    args = parse_args()
    doc_path = Path(args.document).resolve()
    if not doc_path.exists():
        raise SystemExit(f"Document not found: {doc_path}")

    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"[1/5] Parsing document: {doc_path.name}")
    parsed = parse_document(str(doc_path))

    print("[2/5] Running deterministic local checks")
    findings: list[Finding] = []
    findings.extend(run_local_checks(parsed))

    if args.llm:
        print("[3/5] Running LLM engineering review")
        client = build_ollama_client(args.host)
        findings.extend(run_llm_review(client, args.model, parsed))

        if args.vision_model:
            print("[4/5] Running vision review")
            findings.extend(run_vision_review(client, args.vision_model, parsed))
        else:
            print("[4/5] Vision review skipped")
    else:
        print("[3/5] LLM review skipped")
        print("[4/5] Vision review skipped")

    print("[5/5] Validating and writing report")
    findings = validate_and_dedupe(findings, parsed)
    out = output_dir / f"KimiStyle_Report_{doc_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    write_excel(findings, out)

    print(f"Report: {out}")
    print(f"Total findings: {len(findings)}")
    for sev in ["High", "Medium", "Low"]:
        print(f"{sev}: {sum(1 for f in findings if f.severity == sev)}")
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Kimi-style engineering document analyzer")
    parser.add_argument("document", help="Path to DOCX document")
    parser.add_argument("--output-dir", default="kimi_style_analyzer/output")
    parser.add_argument("--llm", action="store_true", help="Enable Ollama LLM passes")
    parser.add_argument("--model", default="gpt-oss:120b")
    parser.add_argument("--vision-model", default="")
    parser.add_argument("--host", default=os.environ.get("OLLAMA_HOST", "https://ollama.com"))
    return parser.parse_args()


def build_ollama_client(host: str) -> Client:
    api_key = os.environ.get("OLLAMA_API_KEY", "").strip()
    if not api_key:
        raise SystemExit("OLLAMA_API_KEY is required when --llm is used")
    return Client(host=host, headers={"Authorization": f"Bearer {api_key}"})


def run_local_checks(parsed: dict) -> list[Finding]:
    findings: list[Finding] = []
    raw_text = collect_text(parsed)

    findings.extend(check_known_typos(raw_text))
    findings.extend(check_table_caption_content(parsed))
    findings.extend(check_datasheet_orphan_notes(raw_text))
    findings.extend(check_engineering_red_flags(raw_text))
    findings.extend(check_table_placeholders(parsed))
    findings.extend(check_decimal_consistency(parsed))
    return findings


def collect_text(parsed: dict) -> str:
    parts = [parsed.get("raw_text", "")]
    for table in parsed.get("tables", []):
        parts.append(table.get("name", ""))
        for row in table.get("rows", []):
            parts.append(" | ".join(row))
    return "\n".join(parts)


def check_known_typos(text: str) -> list[Finding]:
    checks = [
        (r"\bturns ration\b", "Grammar & Spelling", "Typo: 'turns ration' should be 'turns ratio'", "Medium"),
        (r"\bVoit\b", "Grammar & Spelling", "Typo: 'Voit' should likely be 'Vout'", "Medium"),
        (r"characteristic[’']s\b", "Grammar & Spelling", "Typo: 'characteristic's' should be 'characteristics'", "Low"),
        (r"Schematics of its similar", "Grammar & Spelling", "Grammar issue: 'Schematics of its similar' should be 'Schematic of a similar'", "Low"),
        (r"\bT\s*=\s*180\s*kHz\s*=\s*12\.5\s*[uµμ]s", "Units & Calculations", "Unit/formula issue: 'T=180 kHz=12.5 us' is dimensionally invalid", "Medium"),
    ]
    findings = []
    for pattern, category, description, severity in checks:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            evidence = snippet(text, match.start(), match.end())
            findings.append(Finding(category, description, "Document text", severity, evidence, evidence, "local_regex"))
    return findings


def check_table_caption_content(parsed: dict) -> list[Finding]:
    findings: list[Finding] = []
    for table in parsed.get("tables", []):
        name = table.get("name", f"Table {table.get('index', 0) + 1}")
        # The existing parser uses the previous paragraph as a table name. In
        # many Word files that paragraph is not a real caption, so keep this
        # deterministic check conservative and leave fuzzy caption/content
        # review to the LLM table pass.
        if not re.match(r"^\s*Table\s+\d+\b", name, flags=re.IGNORECASE):
            continue
        rows = table.get("rows", [])
        content = " ".join(" ".join(row) for row in rows[:8]).lower()
        caption = name.lower()

        caption_keywords = set(re.findall(r"[a-zA-Z]{4,}", caption))
        content_keywords = set(re.findall(r"[a-zA-Z]{4,}", content))
        generic = {"table", "figure", "value", "rated", "required", "comment", "parameter"}
        caption_keywords -= generic
        content_keywords -= generic

        if len(caption_keywords) < 2 or not content_keywords:
            continue
        overlap = caption_keywords & content_keywords
        if not overlap:
            findings.append(Finding(
                "Cross-Reference Accuracy",
                f"Possible table caption/content mismatch: '{name}'",
                name,
                "High",
                "Caption keywords do not appear in the visible table content. Verify that the caption belongs to this table.",
                "Caption: " + name,
                "local_table_caption",
            ))
    return findings


def check_datasheet_orphan_notes(text: str) -> list[Finding]:
    findings = []
    for match in re.finditer(r"\b(?:Note|NOTE)\s*\d+\b|\^\(Note\s*\d+\)\^|\(\d+\)", text):
        evidence = snippet(text, match.start(), match.end())
        findings.append(Finding(
            "Datasheet Copy Error",
            f"Possible orphan datasheet/reference note: '{match.group(0)}'",
            "Document text",
            "Medium",
            "Verify that the referenced note is defined in this document, not only in the source datasheet.",
            evidence,
            "local_orphan_note",
        ))
    return findings


def check_engineering_red_flags(text: str) -> list[Finding]:
    findings = []
    patterns = [
        (r"650\s*V.*650\s*V|650V.*650V", "Logical Consistency", "Possible zero voltage margin: requirement and device rating both appear to be 650V", "High"),
        (r"1500\s*W.*1350\s*W|1350\s*W.*1500\s*W", "Logical Consistency", "Possible power-rating mismatch: 1500W and 1350W both appear in related context", "High"),
        (r"\bSiC\b", "Terminology Consistency", "Verify SiC terminology against selected semiconductor part numbers", "Medium"),
        (r"inrush", "Logical Consistency", "Inrush current is mentioned; verify that limiting/protection design is specified", "Medium"),
        (r"fuse|circuit breaker", "Logical Consistency", "Input protection is mentioned; verify rating and placement are specified", "Medium"),
    ]
    for pattern, category, description, severity in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE | re.DOTALL):
            evidence = snippet(text, match.start(), match.end())
            findings.append(Finding(category, description, "Document text", severity, evidence, evidence, "local_red_flag"))
            break
    return findings


def check_table_placeholders(parsed: dict) -> list[Finding]:
    findings = []
    placeholder_patterns = re.compile(r"\b(TBD|dummy|placeholder|custom part from TICO 1500W)\b", re.IGNORECASE)
    for table in parsed.get("tables", []):
        name = table.get("name", f"Table {table.get('index', 0) + 1}")
        for r_idx, row in enumerate(table.get("rows", []), start=1):
            row_text = " | ".join(row)
            if placeholder_patterns.search(row_text):
                findings.append(Finding(
                    "Table Quality",
                    "Placeholder or dummy text found in table",
                    f"{name}, row {r_idx}",
                    "High",
                    "Replace placeholder text with actual engineering data or a controlled TBD with owner/date.",
                    row_text[:500],
                    "local_placeholder",
                ))
    return findings


def check_decimal_consistency(parsed: dict) -> list[Finding]:
    findings = []
    number = re.compile(r"^-?\d+\.(\d+)\s*[A-Za-z%Ω℃°]*$")
    for table in parsed.get("tables", []):
        rows = table.get("rows", [])
        if len(rows) < 4:
            continue
        name = table.get("name", f"Table {table.get('index', 0) + 1}")
        max_cols = max((len(row) for row in rows), default=0)
        for col in range(max_cols):
            places: dict[int, list[str]] = {}
            for row in rows[1:]:
                if col >= len(row):
                    continue
                cell = row[col].strip()
                m = number.match(cell)
                if not m:
                    continue
                places.setdefault(len(m.group(1)), []).append(cell)
            if len(places) > 1 and sum(len(v) for v in places.values()) >= 3:
                findings.append(Finding(
                    "Decimal Digit Consistency",
                    "Inconsistent decimal places within a table column",
                    f"{name}, column {col + 1}",
                    "Low",
                    "Standardize precision for values in the same column unless different precision is intentional.",
                    str(places),
                    "local_decimal",
                ))
    return findings


def run_llm_review(client: Client, model: str, parsed: dict) -> list[Finding]:
    findings: list[Finding] = []
    summary = get_document_summary(parsed)

    chunks = get_section_chunks(parsed, max_chars=7000)
    for idx, chunk in enumerate(chunks, start=1):
        prompt = build_chunk_prompt(summary, chunk, idx)
        findings.extend(call_llm_for_findings(client, model, prompt, f"llm_chunk_{idx}"))

    findings.extend(call_llm_for_findings(client, model, build_consistency_prompt(summary), "llm_consistency"))

    for batch_idx, batch in enumerate(batch_tables(parsed.get("tables", []), size=8), start=1):
        prompt = build_table_prompt(batch, batch_idx)
        findings.extend(call_llm_for_findings(client, model, prompt, f"llm_table_{batch_idx}"))

    return findings


def run_vision_review(client: Client, vision_model: str, parsed: dict) -> list[Finding]:
    findings: list[Finding] = []
    images = [img for img in parsed.get("images", []) if img.get("full_b64") and not img.get("is_small")]
    for idx, img in enumerate(images, start=1):
        prompt = build_image_prompt(idx, img)
        try:
            response = client.chat(
                model=vision_model,
                messages=[{"role": "user", "content": prompt, "images": [img["full_b64"]]}],
                options={"temperature": 0.02, "num_predict": 2048},
            )
            reply = response["message"]["content"] if isinstance(response, dict) else response.message.content
            findings.extend(parse_llm_json(reply, f"vision_image_{idx}"))
        except Exception as exc:
            print(f"Vision review failed for image {idx}: {exc}")
    return findings


def build_chunk_prompt(summary: str, chunk: str, idx: int) -> str:
    return f"""You are a senior automotive power-electronics document reviewer.

Find only real, defensible issues. Every issue must include evidence from the text.

Check specifically:
- grammar/spelling and broken subscript/superscript
- terminology inconsistencies: HVDCDC/HV DCDC/HVDC, ACC/USBACC, SiC vs silicon MOSFET
- units and calculations: dimensional errors, impossible negative values, missing units, wrong formulas
- power budget contradictions: 1500W vs 1350W, USB output totals, voltage/current margins
- missing design elements: fuse, breaker, inrush limiting, EMI filter, grounding, thermal model
- missing validation: efficiency criteria, thermal results, PD compliance, CAN FD signal integrity
- connector/pin mapping gaps
- table/figure/caption mismatch
- datasheet copy artifacts and undefined notes

Reject non-issues. Do not report anything that is correct, acceptable, or merely adequate.

Document summary:
{summary[:5000]}

Chunk {idx}:
{chunk}

Return ONLY JSON array:
[
  {{
    "category": "one of: {', '.join(CATEGORIES)}",
    "severity": "High|Medium|Low",
    "description": "short issue title",
    "location": "section/table/figure/page",
    "details": "why it is wrong and what must be checked/fixed",
    "evidence": "exact quote or compact evidence"
  }}
]
"""


def build_consistency_prompt(summary: str) -> str:
    return f"""Review the full document summary for cross-document contradictions.

Prioritize engineering-level issues:
- power budget mismatches
- topology mismatch between diagrams and calculations
- selected component rating equals requirement with no margin
- inconsistent part numbers
- missing SRS mapping coverage
- repeated table-caption/content mismatch pattern
- missing safety/protection/thermal/EMI design evidence

Return ONLY JSON array with category, severity, description, location, details, evidence.

Summary:
{summary[:12000]}
"""


def batch_tables(tables: list[dict], size: int) -> Iterable[list[dict]]:
    for i in range(0, len(tables), size):
        yield tables[i:i + size]


def build_table_prompt(tables: list[dict], batch_idx: int) -> str:
    rendered = []
    for table in tables:
        name = table.get("name", f"Table {table.get('index', 0) + 1}")
        lines = [f"--- {name} ({table.get('num_rows')}x{table.get('num_cols')}) ---"]
        for r_idx, row in enumerate(table.get("rows", [])[:80], start=1):
            lines.append(f"Row {r_idx}: " + " | ".join(cell[:160] for cell in row))
        rendered.append("\n".join(lines))

    return f"""You are reviewing engineering tables. Batch {batch_idx}.

Check:
- caption/content mismatch
- wrong copied table caption
- missing units
- impossible values
- calculation mismatch visible in table
- placeholder/dummy rows
- undefined notes
- missing pin mapping or connector assignment
- missing pass/fail criteria

Do not say a table is truncated unless the visible content explicitly proves it.

Tables:
{chr(10).join(rendered)}

Return ONLY JSON array with category, severity, description, location, details, evidence.
"""


def build_image_prompt(idx: int, image: dict) -> str:
    return f"""Review image {idx} from an engineering document.

Image metadata: {image.get('width')}x{image.get('height')} {image.get('format')}.

Check:
- schematic symbols that do not match part numbers
- duplicate reference designators
- shorted outputs, NC pins connected incorrectly, impossible net labels
- graph axes/units/legends missing
- image content contradicting caption or nearby text
- unreadable/cropped image

Return ONLY JSON array with category, severity, description, location, details, evidence.
"""


def call_llm_for_findings(client: Client, model: str, prompt: str, source: str) -> list[Finding]:
    try:
        response = client.chat(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            options={"temperature": 0.02, "num_predict": 4096},
        )
        reply = response["message"]["content"] if isinstance(response, dict) else response.message.content
        return parse_llm_json(reply, source)
    except Exception as exc:
        print(f"LLM pass failed ({source}): {exc}")
        return []


def parse_llm_json(text: str, source: str) -> list[Finding]:
    text = text.strip()
    if "```json" in text:
        text = text.split("```json", 1)[1].split("```", 1)[0].strip()
    elif "```" in text:
        text = text.split("```", 1)[1].split("```", 1)[0].strip()

    start = text.find("[")
    end = text.rfind("]")
    if start == -1 or end == -1 or end <= start:
        return []

    try:
        payload = json.loads(text[start:end + 1])
    except json.JSONDecodeError:
        return []

    findings = []
    for item in payload:
        if not isinstance(item, dict):
            continue
        findings.append(Finding(
            category=str(item.get("category", "Logical Consistency")),
            description=str(item.get("description", "")).strip(),
            location=str(item.get("location", "-")).strip(),
            severity=str(item.get("severity", "Medium")).strip().title(),
            details=str(item.get("details", "")).strip(),
            evidence=str(item.get("evidence", "")).strip(),
            source=source,
        ))
    return findings


def validate_and_dedupe(findings: list[Finding], parsed: dict) -> list[Finding]:
    valid = []
    seen = set()
    text = collect_text(parsed).lower()

    for finding in findings:
        finding.category = normalize_category(finding.category)
        finding.severity = normalize_severity(finding.severity)

        blob = " ".join([finding.description, finding.details, finding.evidence]).lower()
        if any(phrase in blob for phrase in NON_ERROR_PHRASES):
            continue
        if not finding.description or not finding.details:
            continue
        if finding.source.startswith("llm") and not finding.evidence:
            continue
        if finding.evidence and len(finding.evidence) > 12:
            tokens = [t for t in re.findall(r"[a-zA-Z0-9]{4,}", finding.evidence.lower()) if len(t) > 3]
            if tokens and not any(token in text for token in tokens[:6]):
                continue

        key = (
            finding.category,
            re.sub(r"\W+", " ", finding.description.lower()).strip()[:90],
            re.sub(r"\W+", " ", finding.location.lower()).strip()[:60],
        )
        if key in seen:
            continue
        seen.add(key)
        valid.append(finding)

    valid.sort(key=lambda f: (SEVERITY_ORDER[f.severity], f.category, f.description))
    return valid


def normalize_category(category: str) -> str:
    cleaned = category.strip()
    if cleaned in CATEGORIES:
        return cleaned
    upper_map = {
        "GRAMMAR_SPELLING": "Grammar & Spelling",
        "TERMINOLOGY_CONSISTENCY": "Terminology Consistency",
        "FLOWCHART_DESCRIPTION": "Flowchart ↔ Description",
        "UNITS_CALCULATIONS": "Units & Calculations",
        "FORMATTING_ALIGNMENT": "Formatting & Alignment",
        "SIGNAL_VARIABLE_NAMING": "Signal & Variable Naming",
        "TEST_RESULT_COMPLETENESS": "Test Result Completeness",
        "WAVEFORM_DOCUMENTATION": "Waveform Documentation",
        "CROSS_REFERENCE_ACCURACY": "Cross-Reference Accuracy",
        "LOGICAL_CONSISTENCY": "Logical Consistency",
        "CONNECTOR_PIN_MAPPING": "Connector & Pin Mapping",
        "MEASUREMENT_RESOLUTION": "Measurement Resolution",
        "DECIMAL_CONSISTENCY": "Decimal Digit Consistency",
        "TABLE_QUALITY": "Table Quality",
        "SUBSCRIPT_FORMATTING": "Subscript/Superscript",
        "DATASHEET_COPY_ERROR": "Datasheet Copy Error",
        "TOC_VALIDATION": "TOC & Heading Structure",
    }
    return upper_map.get(cleaned.upper(), "Logical Consistency")


def normalize_severity(severity: str) -> str:
    sev = severity.strip().title()
    if sev == "Critical":
        return "High"
    if sev == "Major":
        return "Medium"
    if sev == "Minor":
        return "Low"
    return sev if sev in SEVERITIES else "Medium"


def write_excel(findings: list[Finding], output_path: Path) -> None:
    rows = []
    for idx, finding in enumerate(findings, start=1):
        rows.append({
            "Error ID": f"ERR-{idx:03d}",
            "Category": finding.category,
            "Icon": finding.icon,
            "Description": finding.description,
            "Location": finding.location,
            "Severity": finding.severity,
            "Details": finding.details,
            "Evidence": finding.evidence,
            "Source": finding.source,
        })

    df = pd.DataFrame(rows, columns=[
        "Error ID", "Category", "Icon", "Description", "Location",
        "Severity", "Details", "Evidence", "Source",
    ])

    summary_rows = []
    for category in sorted(CATEGORIES):
        cat_df = df[df["Category"] == category] if not df.empty else df
        if cat_df.empty:
            continue
        summary_rows.append({
            "Category": category,
            "Icon": CATEGORIES[category],
            "High": int((cat_df["Severity"] == "High").sum()),
            "Medium": int((cat_df["Severity"] == "Medium").sum()),
            "Low": int((cat_df["Severity"] == "Low").sum()),
            "Total": len(cat_df),
        })
    summary = pd.DataFrame(summary_rows)
    if not summary.empty:
        summary.loc[len(summary)] = {
            "Category": "TOTAL",
            "Icon": "",
            "High": int((df["Severity"] == "High").sum()),
            "Medium": int((df["Severity"] == "Medium").sum()),
            "Low": int((df["Severity"] == "Low").sum()),
            "Total": len(df),
        }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Error Details", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        for severity in ["High", "Medium", "Low"]:
            df[df["Severity"] == severity].to_excel(writer, sheet_name=f"{severity} Severity", index=False)
        for category in sorted(df["Category"].unique()) if not df.empty else []:
            safe_name = sanitize_sheet_name(category)
            df[df["Category"] == category].to_excel(writer, sheet_name=safe_name, index=False)

    style_workbook(output_path)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    severity_fill = {
        "High": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "Medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    }

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = [cell.value for cell in ws[1]]
        severity_col = headers.index("Severity") + 1 if "Severity" in headers else None
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if severity_col:
                sev = row[severity_col - 1].value
                if sev in severity_fill:
                    row[severity_col - 1].fill = severity_fill[sev]
                    row[severity_col - 1].font = Font(bold=True)

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
            width = min(max(max_len + 2, 10), 70)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(path)


def sanitize_sheet_name(name: str) -> str:
    safe = re.sub(r"[\\/*?:\[\]]", "_", name)
    safe = safe.replace("↔", "_").replace("&", "_").replace(" ", "_")
    return safe[:31]


def snippet(text: str, start: int, end: int, radius: int = 90) -> str:
    return re.sub(r"\s+", " ", text[max(0, start - radius): min(len(text), end + radius)]).strip()


if __name__ == "__main__":
    raise SystemExit(main())
