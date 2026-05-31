"""
Accuracy scorer: compare DOC-AI findings against a human-reviewed report.

Ground-truth reports vary in layout, so columns are detected by header name
(page / comment(s) / section|item|subject). Matching is intentionally fuzzy
(token overlap + page proximity) because the tool and a human phrase the same
issue differently — the goal is a *directional* precision/recall signal that
makes regressions visible, not a perfect labelling.

Usage:
    # Score a saved findings JSON (from a real LLM run) against a gold report:
    python -m eval.score --gold <human_report.xlsx> --findings <findings.json>

    # Score local (no-LLM) checks on a document against a gold report:
    python -m eval.score --gold <human_report.xlsx> --doc <document.docx> --local-only

    # Score a whole folder of (doc, gold) pairs matched by name:
    python -m eval.score --suite test_data
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

STOPWORDS = {
    "the", "and", "for", "are", "but", "not", "you", "with", "this", "that",
    "from", "have", "has", "was", "were", "should", "would", "could", "into",
    "page", "table", "figure", "section", "comment", "comments", "item", "value",
    "review", "document", "shall", "must", "which", "their", "there", "when",
}

MATCH_THRESHOLD = 0.18  # Jaccard on significant tokens


@dataclass
class GoldItem:
    page: int | None
    location: str
    comment: str
    tokens: set = field(default_factory=set)


def _tokenize(text: str) -> set:
    toks = re.findall(r"[a-zA-Z0-9]{4,}", (text or "").lower())
    return {t for t in toks if t not in STOPWORDS}


def _to_int(val):
    if val is None:
        return None
    m = re.search(r"\d+", str(val))
    return int(m.group()) if m else None


def load_gold(xlsx_path: str) -> list[GoldItem]:
    """Parse a human-reviewed report into ground-truth items."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    items: list[GoldItem] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [r for r in ws.iter_rows(values_only=True) if r and any(c is not None and str(c).strip() for c in r)]
        if not rows:
            continue
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        def find_col(*names):
            for i, h in enumerate(header):
                if h in names:
                    return i
            return None

        col_comment = find_col("comment", "comments")
        if col_comment is None:
            continue  # not a findings sheet
        col_page = find_col("page")
        col_loc = find_col("section", "item", "subject", "location")

        for r in rows[1:]:
            comment = str(r[col_comment]).strip() if col_comment < len(r) and r[col_comment] is not None else ""
            if not comment or comment.lower() == "none":
                continue
            page = _to_int(r[col_page]) if col_page is not None and col_page < len(r) else None
            loc = str(r[col_loc]).strip() if col_loc is not None and col_loc < len(r) and r[col_loc] is not None else ""
            items.append(GoldItem(page=page, location=loc, comment=comment, tokens=_tokenize(comment + " " + loc)))
    wb.close()
    return items


def _pred_tokens(finding: dict) -> set:
    return _tokenize(
        " ".join([
            str(finding.get("comment", "")),
            str(finding.get("evidence", "")),
            str(finding.get("section", "")),
        ])
    )


def _jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def _page_ok(p_pred, p_gold) -> bool:
    if p_pred is None or p_gold is None:
        return True  # page unknown on either side → don't penalize
    return abs(p_pred - p_gold) <= 1


def score(findings: list[dict], gold: list[GoldItem], threshold: float = MATCH_THRESHOLD) -> dict:
    """Compute precision/recall/F1 of `findings` against `gold`."""
    pred = [
        {"tokens": _pred_tokens(f), "page": _to_int(f.get("page")), "raw": f}
        for f in findings
    ]

    matched_gold = set()
    matched_pred = set()
    for gi, g in enumerate(gold):
        best, best_pi = 0.0, None
        for pi, p in enumerate(pred):
            if not _page_ok(p["page"], g.page):
                continue
            j = _jaccard(p["tokens"], g.tokens)
            if j > best:
                best, best_pi = j, pi
        if best >= threshold and best_pi is not None:
            matched_gold.add(gi)
            matched_pred.add(best_pi)

    true_pos = len(matched_gold)
    recall = true_pos / len(gold) if gold else 0.0
    precision = len(matched_pred) / len(pred) if pred else 0.0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0
    return {
        "gold_total": len(gold),
        "predicted_total": len(pred),
        "matched": true_pos,
        "precision": round(precision, 3),
        "recall": round(recall, 3),
        "f1": round(f1, 3),
    }


def _findings_from_doc_local_only(doc_path: str) -> list[dict]:
    from doc_parser import parse_document
    from review_engine import _run_local_checks
    parsed = parse_document(doc_path)
    return _run_local_checks(parsed)


def _print_report(name: str, result: dict):
    print(f"\n=== {name} ===")
    print(f"  gold={result['gold_total']}  predicted={result['predicted_total']}  matched={result['matched']}")
    print(f"  precision={result['precision']}  recall={result['recall']}  f1={result['f1']}")


def main() -> int:
    ap = argparse.ArgumentParser(description="DOC-AI accuracy scorer")
    ap.add_argument("--gold", help="Path to a human-reviewed .xlsx report")
    ap.add_argument("--findings", help="Path to a findings .json (from a real run)")
    ap.add_argument("--doc", help="Path to a .docx to run local-only checks on")
    ap.add_argument("--local-only", action="store_true", help="Run deterministic local checks (no LLM)")
    ap.add_argument("--threshold", type=float, default=MATCH_THRESHOLD)
    args = ap.parse_args()

    if not args.gold:
        ap.error("--gold is required")
    gold = load_gold(args.gold)

    if args.findings:
        with open(args.findings, encoding="utf-8") as fh:
            findings = json.load(fh)
    elif args.doc and args.local_only:
        findings = _findings_from_doc_local_only(args.doc)
    else:
        ap.error("provide --findings <json> OR --doc <docx> --local-only")
        return 2

    result = score(findings, gold, args.threshold)
    _print_report(os.path.basename(args.gold), result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
