"""Tests for batch/suite review: cross-document pass + combined report."""

import os
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from review_engine import review_cross_document
from report_generator import generate_batch_report
from test_engine_pipeline import FakeOllamaClient


class CrossDocTests(unittest.TestCase):
    def test_single_doc_returns_empty(self):
        parsed = {"filename": "a.docx", "raw_text": "x", "sections": [], "tables": [], "images": [], "statistics": {}}
        self.assertEqual(review_cross_document(FakeOllamaClient(), "m", [parsed]), [])

    def test_two_docs_runs(self):
        def _doc(name):
            return {"filename": name, "raw_text": "Scope of this document.", "sections": [],
                    "tables": [], "images": [],
                    "statistics": {"total_words": 3, "total_sections": 0, "total_tables": 0, "total_images": 0}}
        out = review_cross_document(FakeOllamaClient(), "m", [_doc("sdd.docx"), _doc("sctm.xlsx")])
        self.assertIsInstance(out, list)  # mock returns canned findings


class BatchReportTests(unittest.TestCase):
    def test_combined_report_has_all_sheets(self):
        per_doc = [
            {"name": "SDD.docx", "findings": [
                {"category": "GRAMMAR_SPELLING", "severity": "MINOR", "page": "1", "section": "X", "comment": "c", "evidence": "e", "fix": "f"}]},
            {"name": "SCTM.xlsx", "findings": []},
        ]
        cross = [{"category": "TERMINOLOGY_CONSISTENCY", "severity": "MAJOR", "page": "-",
                  "section": "SDD vs SCTM", "comment": "VDD vs VCC", "evidence": "VDD", "fix": "unify"}]
        out = os.path.join(tempfile.gettempdir(), "suite.xlsx")
        generate_batch_report(per_doc, cross, out)
        wb = openpyxl.load_workbook(out)
        self.assertIn("Combined Summary", wb.sheetnames)
        self.assertIn("Cross-Document", wb.sheetnames)
        self.assertTrue(any(s.startswith("SDD") for s in wb.sheetnames))
        os.remove(out)


if __name__ == "__main__":
    unittest.main()
