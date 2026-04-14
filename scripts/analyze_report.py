
import openpyxl
import sys
import json

wb = openpyxl.load_workbook(r'C:\Users\yash badgujar\Downloads\TICO\Review_Report_ACC_Ph3_SDD_8-04-26_Pro_20260413_1354.xlsx')

# ---- Sheet: Review Findings ----
ws = wb['Review Findings']
headers = [cell.value for cell in ws[1]]
findings = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(v is not None for v in row):
        finding = dict(zip(headers, row))
        findings.append(finding)

out = []
for f in findings:
    out.append({
        "no": f.get("No"),
        "page": f.get("Page"),
        "section": f.get("Section", "")[:80] if f.get("Section") else "",
        "severity": f.get("Severity"),
        "category": f.get("Category"),
        "fix_type": f.get("Fix Type"),
        "comment": (f.get("Comment") or "")[:200],
        "fix": (f.get("Fix") or "")[:100],
    })

# Write to file to avoid encoding issues
with open('report_analysis.json', 'w', encoding='utf-8') as fp:
    json.dump({"total": len(findings), "findings": out}, fp, ensure_ascii=False, indent=2)

# Also print summary counts
from collections import Counter
severities = Counter(f.get("Severity") for f in findings)
categories = Counter(f.get("Category") for f in findings)
pages = [f.get("Page") for f in findings if f.get("Page") not in (None, "-", "")]

print(f"TOTAL FINDINGS: {len(findings)}")
print(f"\nSEVERITY BREAKDOWN:")
for k, v in severities.most_common():
    print(f"  {k}: {v}")
print(f"\nCATEGORY BREAKDOWN:")
for k, v in categories.most_common():
    print(f"  {k}: {v}")
print(f"\nPAGE RANGE: {min(pages) if pages else 'N/A'} - {max(pages) if pages else 'N/A'}")
print(f"Pages with findings: {sorted(set(pages))}")
