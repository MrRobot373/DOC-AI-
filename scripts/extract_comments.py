"""
Extract ALL comments from the review files for pattern analysis.
"""
import openpyxl
import os
import sys

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

REVIEW_FILES = [
    r"c:\Users\yash badgujar\Downloads\TICO\Doc_Review_UltraSmall_HDD (1) (1).xlsx",
    r"c:\Users\yash badgujar\Downloads\TICO\Ultrasmall_Ph4_WCCA_SCTM_Review.xlsx",
    r"c:\Users\yash badgujar\Downloads\TICO\Ultrasmall_Ph4_HSIS_Review.xlsx",
]

for filepath in REVIEW_FILES:
    if not os.path.exists(filepath):
        print(f"MISSING: {filepath}")
        continue
    
    fname = os.path.basename(filepath)
    print(f"\n{'='*100}")
    print(f"FILE: {fname}")
    print(f"{'='*100}")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[wb.sheetnames[0]]
        
        # Get all headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers[col] = str(val).strip()
        print(f"Headers: {headers}")
        print(f"Total rows: {ws.max_row}")
        
        # Find key columns
        comment_col = None
        page_col = None
        item_col = None
        rank_col = None
        cat_col = None
        
        for col, name in headers.items():
            nl = name.lower()
            if 'comment' in nl:
                comment_col = col
            elif nl in ('page', 'pg'):
                page_col = col
            elif nl in ('item', 'section', 'chapter'):
                item_col = col
            elif nl in ('rank', 'severity', 'priority'):
                rank_col = col
            elif nl in ('cat', 'category', 'type'):
                cat_col = col
        
        print(f"Comment col: {comment_col}, Page col: {page_col}, Item col: {item_col}, Rank col: {rank_col}, Cat col: {cat_col}")
        
        # Print ALL comments
        for row in range(2, ws.max_row + 1):
            comment = ws.cell(row=row, column=comment_col).value if comment_col else None
            page = ws.cell(row=row, column=page_col).value if page_col else None
            item = ws.cell(row=row, column=item_col).value if item_col else None
            rank = ws.cell(row=row, column=rank_col).value if rank_col else None
            cat = ws.cell(row=row, column=cat_col).value if cat_col else None
            
            if comment:
                comment_str = str(comment).replace('\n', ' ').strip()[:200]
                print(f"\n[#{row-1}] Page:{page} | Item:{item} | Rank:{rank} | Cat:{cat}")
                print(f"  >> {comment_str}")
    except Exception as e:
        print(f"  ERROR: {e}")
