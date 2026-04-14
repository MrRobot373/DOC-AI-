"""
Analyze real TICO review Excel files to extract error patterns and categories.
"""
import openpyxl
import os
import json

REVIEW_FILES = [
    r"c:\Users\yash badgujar\Downloads\TICO\Doc_Review_UltraSmall_HDD (1) (1).xlsx",
    r"c:\Users\yash badgujar\Downloads\TICO\Doc_Review_UltraSmall_HDD (1).xlsx",
    r"c:\Users\yash badgujar\Downloads\TICO\Doc_Review_UltraSmall_HDD.xlsx",
    r"c:\Users\yash badgujar\Downloads\TICO\Ultrasmall_Ph4_HSIS_Review.xlsx",
    r"c:\Users\yash badgujar\Downloads\TICO\Ultrasmall_Ph4_WCCA_SCTM_Review.xlsx",
    r"c:\Users\yash badgujar\Downloads\TICO\Ultrasmall-CAE-CFD-Endo-san-comments.xlsx",
    r"c:\Users\yash badgujar\Downloads\TICO\TICO-ULTRASMALL-PH3-CONCEPT_HSIS_27_02_26 (1).xlsx",
]

for filepath in REVIEW_FILES:
    if not os.path.exists(filepath):
        print(f"MISSING: {filepath}")
        continue
    
    fname = os.path.basename(filepath)
    print(f"\n{'='*80}")
    print(f"FILE: {fname}")
    print(f"{'='*80}")
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ---")
            
            # Print header row
            headers = []
            for col in range(1, min(ws.max_column + 1, 20)):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers.append(f"Col{col}: {str(val)[:50]}")
            print(f"Headers: {headers}")
            
            # Print first 15 data rows to understand content
            for row in range(2, min(ws.max_row + 1, 17)):
                row_data = []
                for col in range(1, min(ws.max_column + 1, 15)):
                    val = ws.cell(row=row, column=col).value
                    if val:
                        row_data.append(f"C{col}:{str(val)[:80]}")
                if row_data:
                    print(f"  Row {row}: {' | '.join(row_data)}")
    except Exception as e:
        print(f"  ERROR: {e}")
