"""
Report Generator Module
Generates Excel (.xlsx) review reports matching TICO's existing format.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from review_engine import REVIEW_CATEGORIES, SEVERITY_LEVELS


def generate_excel_report(findings, doc_filename, output_path):
    """
    Generate an Excel review report matching the TICO format.

    Columns: No | Page | Section | Comment | Category | Severity | Date | Status
    Also includes a Summary sheet.
    """
    wb = Workbook()

    # ---- Sheet 1: Review Findings ----
    ws = wb.active
    ws.title = "Review Findings"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    severity_fills = {
        "CRITICAL": PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid"),
        "MAJOR": PatternFill(start_color="ffe0b2", end_color="ffe0b2", fill_type="solid"),
        "MINOR": PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid"),
    }

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    body_font = Font(name="Calibri", size=10)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # Status color fills
    status_fills = {
        "OPEN": PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid"),
        "WORKING": PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid"),
        "CLOSED": PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid"),
        "IGNORE": PatternFill(start_color="e0e0e0", end_color="e0e0e0", fill_type="solid"),
        "N/A": PatternFill(start_color="e0e0e0", end_color="e0e0e0", fill_type="solid"),
    }

    # Headers
    headers = ["No", "Page", "Section", "Comment", "Fix", "Category", "Severity", "Date", "Status"]
    col_widths = [6, 8, 15, 50, 50, 22, 12, 14, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Data rows
    review_date = datetime.now().strftime("%Y-%m-%d")

    for row_idx, finding in enumerate(findings, 2):
        cat_info = REVIEW_CATEGORIES.get(finding["category"], {})
        cat_name = cat_info.get("name", finding["category"])
        cat_icon = cat_info.get("icon", "")

        sev_label = SEVERITY_LEVELS.get(finding["severity"], {}).get("label", finding["severity"])
        sev_fill = severity_fills.get(finding["severity"], PatternFill())

        row_data = [
            finding.get("id", row_idx - 1),
            finding.get("page", "-"),
            finding.get("section", "-"),
            finding.get("comment", ""),
            finding.get("fix", ""),
            f"{cat_icon} {cat_name}",
            sev_label,
            review_date,
            "OPEN",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = border
            if col_idx in (1, 2, 7, 8, 9):
                cell.alignment = center_align
            else:
                cell.alignment = wrap_align

            # Apply severity color to severity column
            if col_idx == 7:
                cell.fill = sev_fill
                cell.font = Font(name="Calibri", size=10, bold=True)

            # Apply status color to status column
            if col_idx == 9:
                cell.fill = status_fills.get(str(value), PatternFill())
                cell.font = Font(name="Calibri", size=10, bold=True)

    # Add data validation dropdown for Status column
    status_col_letter = get_column_letter(9)  # Column I = Status
    dv = DataValidation(
        type="list",
        formula1='"OPEN,WORKING,CLOSED,IGNORE,N/A"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid Status",
        error="Please select: OPEN, WORKING, CLOSED, IGNORE, or N/A",
    )
    dv.sqref = f"{status_col_letter}2:{status_col_letter}{len(findings) + 1}"
    ws.add_data_validation(dv)

    # Row height for readability
    for row_idx in range(2, len(findings) + 2):
        ws.row_dimensions[row_idx].height = 45

    # ---- Sheet 2: Summary ----
    ws2 = wb.create_sheet("Summary")

    # Title
    ws2.merge_cells("A1:D1")
    title_cell = ws2.cell(row=1, column=1, value=f"Review Summary — {doc_filename}")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1a1a2e")
    title_cell.alignment = Alignment(horizontal="center")

    ws2.cell(row=2, column=1, value=f"Generated: {review_date}")
    ws2.cell(row=2, column=1).font = Font(name="Calibri", size=10, italic=True, color="666666")

    # Severity breakdown
    ws2.cell(row=4, column=1, value="Severity Breakdown").font = Font(name="Calibri", bold=True, size=12)

    severity_counts = {}
    for f in findings:
        sev = f["severity"]
        severity_counts[sev] = severity_counts.get(sev, 0) + 1

    row = 5
    summary_headers = ["Severity", "Count", "Percentage"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    total = len(findings) if findings else 1
    for sev in ["CRITICAL", "MAJOR", "MINOR"]:
        row += 1
        count = severity_counts.get(sev, 0)
        pct = f"{count / total * 100:.1f}%"
        label = SEVERITY_LEVELS[sev]["label"]

        ws2.cell(row=row, column=1, value=label).border = border
        ws2.cell(row=row, column=1).fill = severity_fills.get(sev, PatternFill())
        ws2.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10)
        ws2.cell(row=row, column=2, value=count).border = border
        ws2.cell(row=row, column=2).alignment = center_align
        ws2.cell(row=row, column=3, value=pct).border = border
        ws2.cell(row=row, column=3).alignment = center_align

    row += 1
    ws2.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True, size=10)
    ws2.cell(row=row, column=1).border = border
    ws2.cell(row=row, column=2, value=len(findings)).border = border
    ws2.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=10)
    ws2.cell(row=row, column=2).alignment = center_align

    # Category breakdown
    row += 2
    ws2.cell(row=row, column=1, value="Category Breakdown").font = Font(name="Calibri", bold=True, size=12)

    row += 1
    cat_headers = ["Category", "Count", "Critical", "Major"]
    for col_idx, h in enumerate(cat_headers, 1):
        cell = ws2.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    cat_counts = {}
    cat_critical = {}
    cat_major = {}
    for f in findings:
        cat = f["category"]
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
        if f["severity"] == "CRITICAL":
            cat_critical[cat] = cat_critical.get(cat, 0) + 1
        if f["severity"] == "MAJOR":
            cat_major[cat] = cat_major.get(cat, 0) + 1

    for cat_id, count in sorted(cat_counts.items(), key=lambda x: x[1], reverse=True):
        row += 1
        cat_info = REVIEW_CATEGORIES.get(cat_id, {})
        cat_name = f"{cat_info.get('icon', '')} {cat_info.get('name', cat_id)}"
        ws2.cell(row=row, column=1, value=cat_name).border = border
        ws2.cell(row=row, column=2, value=count).border = border
        ws2.cell(row=row, column=2).alignment = center_align
        ws2.cell(row=row, column=3, value=cat_critical.get(cat_id, 0)).border = border
        ws2.cell(row=row, column=3).alignment = center_align
        ws2.cell(row=row, column=4, value=cat_major.get(cat_id, 0)).border = border
        ws2.cell(row=row, column=4).alignment = center_align

    # Set column widths for summary
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12

    # Save
    wb.save(output_path)
    return output_path
