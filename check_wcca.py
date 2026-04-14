"""Extract comments from WCCA review which has different structure."""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

filepath = r"c:\Users\yash badgujar\Downloads\TICO\Ultrasmall_Ph4_WCCA_SCTM_Review.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)

for sheet_name in wb.sheetnames[:3]:  # First 3 sheets
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ===")
    # Print first 25 rows to understand structure
    for row in range(1, min(ws.max_row + 1, 26)):
        row_data = []
        for col in range(1, min(ws.max_column + 1, 12)):
            val = ws.cell(row=row, column=col).value
            if val:
                s = str(val).replace('\n', ' ')[:60]
                row_data.append(f"C{col}:{s}")
        if row_data:
            print(f"  R{row}: {' | '.join(row_data)}")
